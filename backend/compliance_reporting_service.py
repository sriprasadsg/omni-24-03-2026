import csv
import io
import os
import re
from datetime import datetime
from database import get_database
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table as PDFTable, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ── Scoring helpers ───────────────────────────────────────────────────────────

def _score_status(status: str) -> str:
    """Normalise asset_compliance status to Compliant / Warning / Non-Compliant."""
    s = (status or "").strip()
    if s in ("Compliant", "Pass", "Passed", "Implemented", "pass", "passed"):
        return "Compliant"
    if s in ("Non-Compliant", "Fail", "Failed", "Not Implemented", "fail", "failed"):
        return "Non-Compliant"
    return "Warning"


def _compliance_score(counts: dict) -> float:
    total = counts.get("Compliant", 0) + counts.get("Warning", 0) + counts.get("Non-Compliant", 0)
    if total == 0:
        return 0.0
    return round(counts.get("Compliant", 0) / total * 100, 1)


def _overall_verdict(score: float) -> str:
    if score >= 80:
        return "Compliant"
    if score >= 50:
        return "Partially Compliant"
    return "Non-Compliant"


# ── Data fetching ─────────────────────────────────────────────────────────────

def _flatten_evidence(evidence_list: list) -> dict:
    """
    Extract display fields from a list of evidence/artifact records.
    Handles both asset_compliance evidence entries and compliance_artifact records.
    Returns a dict with: names, urls, descriptions, uploaded_ats, statuses, count.
    """
    names, urls, descs, dates, statuses = [], [], [], [], []
    seen_ids: set = set()
    for e in evidence_list:
        eid = e.get("id") or e.get("url") or e.get("name", "")
        if eid in seen_ids:
            continue
        seen_ids.add(eid)
        name = e.get("name") or e.get("filename") or ""
        url  = e.get("url") or ""
        desc = e.get("description") or ""
        date = (e.get("uploadedAt") or e.get("uploaded_at") or
                e.get("date") or e.get("lastUpdated") or "")
        # Truncate ISO timestamp to date only
        if date and "T" in date:
            date = date[:10]
        status = e.get("status") or ""
        if name:
            names.append(name)
        if url:
            urls.append(url)
        if desc:
            descs.append(desc)
        if date:
            dates.append(date)
        if status:
            statuses.append(status)
    return {
        "names":        names,
        "urls":         urls,
        "descriptions": descs,
        "uploaded_ats": dates,
        "statuses":     statuses,
        "count":        len(seen_ids),
    }


async def _build_report_data(framework_id: str):
    """
    Returns:
        framework      – the framework document
        asset_summary  – list of {assetId, hostname, counts, score, verdict}
        control_rows   – list of dicts with control + evidence columns
    """
    db = get_database()

    framework = await db.compliance_frameworks.find_one({"id": framework_id})
    if not framework:
        raise ValueError(f"Framework '{framework_id}' not found")

    controls = framework.get("controls", [])
    control_ids = [c["id"] for c in controls]

    # ── Fetch asset_compliance docs ───────────────────────────────────────────
    cursor = db.asset_compliance.find({"controlId": {"$in": control_ids}})
    ac_docs = await cursor.to_list(length=10000)

    # ── Fetch standalone artifacts linked to these controls ───────────────────
    art_cursor = db.compliance_artifacts.find({"control_ids": {"$in": control_ids}})
    artifact_docs = await art_cursor.to_list(length=5000)
    # Map control_id -> list of artifacts
    artifact_by_ctrl: dict = {}
    for art in artifact_docs:
        for cid in (art.get("control_ids") or []):
            artifact_by_ctrl.setdefault(cid, []).append(art)

    # ── Fetch asset hostnames ─────────────────────────────────────────────────
    asset_ids = list({d.get("assetId") for d in ac_docs if d.get("assetId")})
    asset_cursor = db.assets.find({"id": {"$in": asset_ids}}, {"id": 1, "hostname": 1, "_id": 0})
    asset_docs = await asset_cursor.to_list(length=1000)
    hostname_map = {a["id"]: a.get("hostname", a["id"]) for a in asset_docs}

    # ── Build per-asset aggregates ────────────────────────────────────────────
    asset_counts: dict = {}
    for doc in ac_docs:
        aid  = doc.get("assetId", "unknown")
        norm = _score_status(doc.get("status", ""))
        if aid not in asset_counts:
            asset_counts[aid] = {"Compliant": 0, "Warning": 0, "Non-Compliant": 0}
        asset_counts[aid][norm] += 1

    asset_summary = []
    for aid, counts in sorted(asset_counts.items()):
        score = _compliance_score(counts)
        asset_summary.append({
            "Asset ID":       aid,
            "Hostname":       hostname_map.get(aid, aid),
            "Total Controls": sum(counts.values()),
            "Compliant":      counts["Compliant"],
            "Warning":        counts["Warning"],
            "Non-Compliant":  counts["Non-Compliant"],
            "Score (%)":      score,
            "Overall Status": _overall_verdict(score),
        })

    # ── Group asset_compliance docs by (controlId, assetId) ──────────────────
    ac_by_key: dict = {}
    for doc in ac_docs:
        key = (doc.get("controlId"), doc.get("assetId", "unknown"))
        ac_by_key[key] = doc

    # ── Build per-control rows ────────────────────────────────────────────────
    control_rows = []
    for ctrl in controls:
        cid          = ctrl["id"]
        cname        = ctrl.get("name", "")
        ccat         = ctrl.get("category", "")
        ctrl_status  = ctrl.get("status", "Not Implemented")
        last_reviewed = ctrl.get("lastReviewed", "")

        # Standalone artifacts for this control (not asset-scoped)
        standalone_arts = artifact_by_ctrl.get(cid, [])

        matching_assets = [(k[1], v) for k, v in ac_by_key.items() if k[0] == cid]

        if not matching_assets:
            # No asset data — show standalone artifacts if any
            ev = _flatten_evidence(standalone_arts)
            control_rows.append({
                "Control ID":       cid,
                "Control Name":     cname,
                "Category":         ccat,
                "Control Status":   ctrl_status,
                "Asset ID":         "—",
                "Hostname":         "—",
                "Asset Status":     "—",
                "Evidence Count":   ev["count"],
                "Evidence Names":   ", ".join(ev["names"]) if ev["names"] else "None",
                "Evidence URLs":    ", ".join(ev["urls"])  if ev["urls"]  else "—",
                "Evidence Dates":   ", ".join(ev["uploaded_ats"]) if ev["uploaded_ats"] else "—",
                "Evidence Desc":    "; ".join(ev["descriptions"])  if ev["descriptions"]  else "—",
                "Last Reviewed":    last_reviewed,
            })
            continue

        for aid, doc in matching_assets:
            # Merge asset evidence + standalone artifacts
            asset_evidence  = doc.get("evidence", [])
            merged_evidence = asset_evidence + [
                a for a in standalone_arts
                if not any(ae.get("id") == a.get("id") for ae in asset_evidence)
            ]
            ev = _flatten_evidence(merged_evidence)

            control_rows.append({
                "Control ID":       cid,
                "Control Name":     cname,
                "Category":         ccat,
                "Control Status":   ctrl_status,
                "Asset ID":         aid,
                "Hostname":         hostname_map.get(aid, aid),
                "Asset Status":     _score_status(doc.get("status", "")),
                "Evidence Count":   ev["count"],
                "Evidence Names":   ", ".join(ev["names"]) if ev["names"] else "None",
                "Evidence URLs":    ", ".join(ev["urls"])  if ev["urls"]  else "—",
                "Evidence Dates":   ", ".join(ev["uploaded_ats"]) if ev["uploaded_ats"] else "—",
                "Evidence Desc":    "; ".join(ev["descriptions"])  if ev["descriptions"]  else "—",
                "Last Reviewed":    last_reviewed,
            })

    return framework, asset_summary, control_rows


# ── CSV ───────────────────────────────────────────────────────────────────────

async def _generate_csv(framework_id: str, reports_dir: str) -> dict:
    framework, asset_summary, control_rows = await _build_report_data(framework_id)
    fw_name = framework.get("name", framework_id)

    output = io.StringIO()
    w = csv.writer(output)

    # Header
    w.writerow([f"Compliance Report: {fw_name}"])
    w.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    w.writerow([])

    # Asset summary section
    w.writerow(["ASSET COMPLIANCE SUMMARY"])
    if asset_summary:
        w.writerow(list(asset_summary[0].keys()))
        for row in asset_summary:
            w.writerow(list(row.values()))
    else:
        w.writerow(["No asset compliance data available"])
    w.writerow([])

    # Control detail section
    w.writerow(["CONTROL DETAILS WITH EVIDENCE"])
    if control_rows:
        w.writerow(list(control_rows[0].keys()))
        for row in control_rows:
            w.writerow(list(row.values()))

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"compliance_report_{framework_id}_{timestamp}.csv"
    filepath = os.path.join(reports_dir, filename)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        f.write(output.getvalue())

    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now().isoformat(),
        "rowCount": len(control_rows),
    }


# ── Excel ─────────────────────────────────────────────────────────────────────

_STATUS_FILLS = {
    "Compliant":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Partially Compliant": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Non-Compliant":       PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Warning":             PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Implemented":         PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Not Implemented":     PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "In Progress":         PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}
_STATUS_FONTS = {
    "Compliant":           Font(color="006100"),
    "Partially Compliant": Font(color="9C5700"),
    "Non-Compliant":       Font(color="9C0006"),
    "Warning":             Font(color="9C5700"),
    "Implemented":         Font(color="006100"),
    "Not Implemented":     Font(color="9C0006"),
    "In Progress":         Font(color="9C5700"),
}
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_SECTION_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)


def _xl_header_row(ws, headers):
    ws.append(headers)
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)


async def _generate_excel(framework_id: str, reports_dir: str) -> dict:
    framework, asset_summary, control_rows = await _build_report_data(framework_id)
    fw_name = framework.get("name", framework_id)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Asset Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Asset Summary"

    # Title
    ws1.append([f"Compliance Report: {fw_name}"])
    ws1["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws1.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws1.append([])

    ws1.append(["ASSET COMPLIANCE SUMMARY"])
    ws1[f"A{ws1.max_row}"].fill = _SECTION_FILL
    ws1[f"A{ws1.max_row}"].font = _SECTION_FONT
    ws1.append([])

    if asset_summary:
        headers = list(asset_summary[0].keys())
        _xl_header_row(ws1, headers)
        status_col = headers.index("Overall Status") + 1
        score_col  = headers.index("Score (%)") + 1
        for row in asset_summary:
            ws1.append(list(row.values()))
            r = ws1.max_row
            for c in range(1, len(headers) + 1):
                ws1.cell(r, c).border = _THIN_BORDER
                ws1.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            # Colour Overall Status cell
            verdict = row.get("Overall Status", "")
            fill = _STATUS_FILLS.get(verdict)
            font = _STATUS_FONTS.get(verdict)
            if fill:
                ws1.cell(r, status_col).fill = fill
            if font:
                ws1.cell(r, status_col).font = font
            # Bold score
            ws1.cell(r, score_col).font = Font(bold=True)
    _xl_auto_width(ws1)

    # ── Sheet 2: Control Details ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Control Details")
    ws2.append([f"Control Details with Evidence — {fw_name}"])
    ws2["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws2.append([])

    if control_rows:
        headers2 = list(control_rows[0].keys())
        _xl_header_row(ws2, headers2)
        ctrl_status_col  = headers2.index("Control Status") + 1
        asset_status_col = headers2.index("Asset Status") + 1
        url_col = headers2.index("Evidence URLs") + 1 if "Evidence URLs" in headers2 else None
        for row in control_rows:
            ws2.append(list(row.values()))
            r = ws2.max_row
            for c in range(1, len(headers2) + 1):
                ws2.cell(r, c).border = _THIN_BORDER
                ws2.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            for col_idx, key in [(ctrl_status_col, "Control Status"), (asset_status_col, "Asset Status")]:
                val = row.get(key, "")
                fill = _STATUS_FILLS.get(val)
                font = _STATUS_FONTS.get(val)
                if fill:
                    ws2.cell(r, col_idx).fill = fill
                if font:
                    ws2.cell(r, col_idx).font = font
            # Make first URL clickable
            if url_col:
                raw_urls = row.get("Evidence URLs", "—")
                if raw_urls and raw_urls != "—":
                    first_url = raw_urls.split(",")[0].strip()
                    if first_url.startswith("http") or first_url.startswith("/"):
                        cell = ws2.cell(r, url_col)
                        cell.hyperlink = first_url
                        cell.font = Font(color="0563C1", underline="single")
    _xl_auto_width(ws2)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"compliance_report_{framework_id}_{timestamp}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)

    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now().isoformat(),
        "rowCount": len(control_rows),
    }


# ── PDF ───────────────────────────────────────────────────────────────────────

_PDF_STATUS_COLORS = {
    "Compliant":           colors.HexColor("#C6EFCE"),
    "Partially Compliant": colors.HexColor("#FFEB9C"),
    "Non-Compliant":       colors.HexColor("#FFC7CE"),
    "Warning":             colors.HexColor("#FFEB9C"),
    "Implemented":         colors.HexColor("#C6EFCE"),
    "Not Implemented":     colors.HexColor("#FFC7CE"),
    "In Progress":         colors.HexColor("#FFEB9C"),
    "—":                   colors.white,
}


def _find_status_rows(rows_data, col_idx: int):
    """Return list of (row_index, bg_color) for coloring status cells."""
    cmds = []
    for i, row in enumerate(rows_data[1:], start=1):  # skip header
        val = str(row[col_idx]) if col_idx < len(row) else ""
        bg = _PDF_STATUS_COLORS.get(val, colors.white)
        if bg != colors.white:
            cmds.append(("BACKGROUND", (col_idx, i), (col_idx, i), bg))
    return cmds


async def _generate_pdf(framework_id: str, reports_dir: str) -> dict:
    framework, asset_summary, control_rows = await _build_report_data(framework_id)
    fw_name = framework.get("name", framework_id)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"compliance_report_{framework_id}_{timestamp}.pdf"
    filepath = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.5 * inch,  bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], fontSize=16,
                                  textColor=colors.HexColor("#1F3864"), spaceAfter=4)
    sub_style   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                                  textColor=colors.HexColor("#555555"), spaceAfter=12)
    section_style = ParagraphStyle("SEC", parent=styles["Heading2"], fontSize=11,
                                    textColor=colors.HexColor("#1F3864"), spaceBefore=14, spaceAfter=6)
    cell_style  = ParagraphStyle("C", parent=styles["Normal"], fontSize=7, leading=9)
    hdr_style   = ParagraphStyle("H", parent=styles["Normal"], fontSize=7, leading=9,
                                  fontName="Helvetica-Bold", textColor=colors.whitesmoke)

    page_w = landscape(letter)[0] - inch  # usable width

    def make_table(headers, rows_plain, col_w_hints):
        """Build a ReportLab Table with wrapped Paragraph cells."""
        total_hint = sum(col_w_hints)
        scale = page_w / total_hint if total_hint > page_w else 1.0
        col_ws = [w * scale * inch for w in col_w_hints]

        hdr_row = [Paragraph(h, hdr_style) for h in headers]
        table_data = [hdr_row]
        for row in rows_plain:
            table_data.append([Paragraph(str(v), cell_style) for v in row])

        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#366092")),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E0")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
        ]
        return PDFTable(table_data, colWidths=col_ws, repeatRows=1), style_cmds

    elements = []
    elements.append(Paragraph(f"{fw_name.upper()} Compliance Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp; "
        f"Controls: {len(framework.get('controls', []))} &nbsp;|&nbsp; "
        f"Assets: {len(asset_summary)}",
        sub_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E0")))

    # ── Asset Summary table ────────────────────────────────────────────────────
    elements.append(Paragraph("Asset Compliance Summary", section_style))

    if asset_summary:
        sum_headers = list(asset_summary[0].keys())
        sum_rows = [list(r.values()) for r in asset_summary]
        # col widths (inches)
        sum_widths = [1.4, 1.2, 0.9, 0.9, 0.8, 1.0, 0.8, 1.2]
        sum_widths = sum_widths[:len(sum_headers)]

        t, base_cmds = make_table(sum_headers, sum_rows, sum_widths)
        # Colour verdict & score columns
        verdict_idx = sum_headers.index("Overall Status")
        status_cmds = _find_status_rows(
            [sum_headers] + sum_rows, verdict_idx
        )
        t.setStyle(TableStyle(base_cmds + status_cmds))
        elements.append(t)
    else:
        elements.append(Paragraph("No asset compliance data available for this framework.", styles["Normal"]))

    elements.append(Spacer(1, 0.2 * inch))

    # ── Control Details table ──────────────────────────────────────────────────
    elements.append(Paragraph("Control Details with Evidence", section_style))

    if control_rows:
        det_headers = list(control_rows[0].keys())
        det_rows    = [list(r.values()) for r in control_rows]
        # widths: ID, Name, Cat, CtrlStatus, AssetID, Hostname, AssetStatus,
        #         EvCount, EvNames, EvURLs, EvDates, EvDesc, LastReviewed
        det_widths  = [0.7, 1.6, 0.9, 0.9, 0.8, 0.9, 0.9, 0.5, 1.6, 1.4, 0.8, 1.2, 0.8]
        det_widths  = det_widths[:len(det_headers)]

        t2, base_cmds2 = make_table(det_headers, det_rows, det_widths)
        ctrl_status_idx  = det_headers.index("Control Status")
        asset_status_idx = det_headers.index("Asset Status")
        color_cmds = (
            _find_status_rows([det_headers] + det_rows, ctrl_status_idx) +
            _find_status_rows([det_headers] + det_rows, asset_status_idx)
        )
        t2.setStyle(TableStyle(base_cmds2 + color_cmds))
        elements.append(t2)

    doc.build(elements)
    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now().isoformat(),
        "rowCount": len(control_rows),
    }


# ── All-frameworks CSV ────────────────────────────────────────────────────────

async def _generate_all_csv(reports_dir: str) -> dict:
    db = get_database()
    frameworks = await db.compliance_frameworks.find({}, {"id": 1, "name": 1}).to_list(length=100)
    if not frameworks:
        raise ValueError("No compliance frameworks found")

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["All Compliance Frameworks — Combined Report"])
    w.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    w.writerow([])

    total_done = 0
    for fw in frameworks:
        fw_id = fw.get("id", "")
        fw_name = fw.get("name", fw_id)
        try:
            _, asset_summary, control_rows = await _build_report_data(fw_id)
        except Exception:
            continue
        total_done += 1

        w.writerow([f"=== FRAMEWORK: {fw_name} ==="])
        w.writerow(["ASSET COMPLIANCE SUMMARY"])
        if asset_summary:
            w.writerow(list(asset_summary[0].keys()))
            for row in asset_summary:
                w.writerow(list(row.values()))
        else:
            w.writerow(["No asset compliance data available"])
        w.writerow([])

        w.writerow(["CONTROL DETAILS WITH EVIDENCE"])
        if control_rows:
            w.writerow(list(control_rows[0].keys()))
            for row in control_rows:
                w.writerow(list(row.values()))
        else:
            w.writerow(["No control data available"])
        w.writerow([])
        w.writerow([])

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"all_compliance_report_{timestamp}.csv"
    filepath = os.path.join(reports_dir, filename)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        f.write(output.getvalue())

    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now().isoformat(),
        "frameworkCount": total_done,
    }


# ── All-frameworks Excel ───────────────────────────────────────────────────────

async def _generate_all_excel(reports_dir: str) -> dict:
    db = get_database()
    frameworks = await db.compliance_frameworks.find({}, {"id": 1, "name": 1}).to_list(length=100)
    if not frameworks:
        raise ValueError("No compliance frameworks found")

    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "Overview"

    ov.append(["All Compliance Frameworks — Overview"])
    ov["A1"].font = Font(bold=True, size=14, color="1F3864")
    ov.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ov.append([])
    ov.append(["FRAMEWORK SUMMARY"])
    ov[f"A{ov.max_row}"].fill = _SECTION_FILL
    ov[f"A{ov.max_row}"].font = _SECTION_FONT
    ov.append([])

    ov_hdrs = ["Framework", "Total Controls", "Assets Evaluated",
               "Avg Score (%)", "Compliant Assets", "Non-Compliant Assets", "Overall Status"]
    _xl_header_row(ov, ov_hdrs)
    ov_status_col = ov_hdrs.index("Overall Status") + 1

    total_done = 0
    for fw in frameworks:
        fw_id = fw.get("id", "")
        fw_name = fw.get("name", fw_id)
        try:
            _, asset_summary, control_rows = await _build_report_data(fw_id)
        except Exception:
            continue
        total_done += 1

        total_controls = len({r["Control ID"] for r in control_rows})
        total_assets = len(asset_summary)
        avg_score = round(sum(a["Score (%)"] for a in asset_summary) / total_assets, 1) if total_assets else 0.0
        compliant_ct = sum(1 for a in asset_summary if a["Overall Status"] == "Compliant")
        non_compliant_ct = sum(1 for a in asset_summary if a["Overall Status"] == "Non-Compliant")
        overall = _overall_verdict(avg_score)

        ov.append([fw_name, total_controls, total_assets, avg_score, compliant_ct, non_compliant_ct, overall])
        r = ov.max_row
        for c in range(1, len(ov_hdrs) + 1):
            ov.cell(r, c).border = _THIN_BORDER
            ov.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        if _STATUS_FILLS.get(overall):
            ov.cell(r, ov_status_col).fill = _STATUS_FILLS[overall]
        if _STATUS_FONTS.get(overall):
            ov.cell(r, ov_status_col).font = _STATUS_FONTS[overall]

        short = re.sub(r'[\\/?*\[\]:]', '-', fw_name)[:24].strip()

        # Asset Summary sheet
        ws_a = wb.create_sheet(f"{short} Assets")
        ws_a.append([f"Asset Summary — {fw_name}"])
        ws_a["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws_a.append([])
        if asset_summary:
            hdrs_a = list(asset_summary[0].keys())
            _xl_header_row(ws_a, hdrs_a)
            sc_a = hdrs_a.index("Overall Status") + 1
            for row in asset_summary:
                ws_a.append(list(row.values()))
                rn = ws_a.max_row
                for c in range(1, len(hdrs_a) + 1):
                    ws_a.cell(rn, c).border = _THIN_BORDER
                    ws_a.cell(rn, c).alignment = Alignment(vertical="top", wrap_text=True)
                vd = row.get("Overall Status", "")
                if _STATUS_FILLS.get(vd): ws_a.cell(rn, sc_a).fill = _STATUS_FILLS[vd]
                if _STATUS_FONTS.get(vd): ws_a.cell(rn, sc_a).font = _STATUS_FONTS[vd]
        _xl_auto_width(ws_a)

        # Control Details sheet
        ws_c = wb.create_sheet(f"{short} Controls")
        ws_c.append([f"Control Details — {fw_name}"])
        ws_c["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws_c.append([])
        if control_rows:
            hdrs_c = list(control_rows[0].keys())
            _xl_header_row(ws_c, hdrs_c)
            cs_c  = hdrs_c.index("Control Status") + 1
            as_c  = hdrs_c.index("Asset Status") + 1
            url_c = hdrs_c.index("Evidence URLs") + 1 if "Evidence URLs" in hdrs_c else None
            for row in control_rows:
                ws_c.append(list(row.values()))
                rn = ws_c.max_row
                for c in range(1, len(hdrs_c) + 1):
                    ws_c.cell(rn, c).border = _THIN_BORDER
                    ws_c.cell(rn, c).alignment = Alignment(vertical="top", wrap_text=True)
                for col_idx, key in [(cs_c, "Control Status"), (as_c, "Asset Status")]:
                    val = row.get(key, "")
                    if _STATUS_FILLS.get(val): ws_c.cell(rn, col_idx).fill = _STATUS_FILLS[val]
                    if _STATUS_FONTS.get(val): ws_c.cell(rn, col_idx).font = _STATUS_FONTS[val]
                if url_c:
                    raw_urls = row.get("Evidence URLs", "—")
                    if raw_urls and raw_urls != "—":
                        first_url = raw_urls.split(",")[0].strip()
                        if first_url.startswith("http") or first_url.startswith("/"):
                            cell = ws_c.cell(rn, url_c)
                            cell.hyperlink = first_url
                            cell.font = Font(color="0563C1", underline="single")
        _xl_auto_width(ws_c)

    _xl_auto_width(ov)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"all_compliance_report_{timestamp}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)

    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now().isoformat(),
        "frameworkCount": total_done,
    }


# ── Service class (backwards-compatible API) ──────────────────────────────────

class ComplianceReportingService:
    def __init__(self):
        self.reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "reports")
        os.makedirs(self.reports_dir, exist_ok=True)

    async def generate_report(self, tenant_id: str, framework_id: str) -> dict:
        return await _generate_csv(framework_id, self.reports_dir)

    async def generate_excel_report(self, tenant_id: str, framework_id: str) -> dict:
        return await _generate_excel(framework_id, self.reports_dir)

    async def generate_pdf_report(self, tenant_id: str, framework_id: str) -> dict:
        return await _generate_pdf(framework_id, self.reports_dir)

    async def generate_all_csv_report(self, tenant_id: str) -> dict:
        return await _generate_all_csv(self.reports_dir)

    async def generate_all_excel_report(self, tenant_id: str) -> dict:
        return await _generate_all_excel(self.reports_dir)


compliance_reporting_service = ComplianceReportingService()
