"""
Compliance reporting: Excel (XLSX) generation for single and all frameworks.
"""

import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from compliance_reporting_data import _build_report_data, _overall_verdict, _sanitize_cell

# ── Shared style constants ────────────────────────────────────────────────────

_STATUS_FILLS = {
    "Compliant":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Partially Compliant": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Non-Compliant":       PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Warning":             PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Implemented":         PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Not Implemented":     PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "In Progress":         PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}
_STATUS_FONTS = {
    "Compliant":           Font(color="006100"),
    "Partially Compliant": Font(color="9C5700"),
    "Non-Compliant":       Font(color="9C0006"),
    "Warning":             Font(color="9C5700"),
    "Implemented":         Font(color="006100"),
    "Not Implemented":     Font(color="9C0006"),
    "In Progress":         Font(color="9C5700"),
}
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_SECTION_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)


# ── Worksheet helpers ─────────────────────────────────────────────────────────

def _xl_header_row(ws, headers):
    ws.append(headers)
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)


def _apply_status_colors(ws, row_num, col_idx, val):
    if _STATUS_FILLS.get(val):
        ws.cell(row_num, col_idx).fill = _STATUS_FILLS[val]
    if _STATUS_FONTS.get(val):
        ws.cell(row_num, col_idx).font = _STATUS_FONTS[val]


def _apply_url_hyperlink(ws, row_num, url_col, raw_urls):
    if url_col and raw_urls and raw_urls != "—":
        first_url = raw_urls.split(",")[0].strip()
        if first_url.startswith("http") or first_url.startswith("/"):
            cell = ws.cell(row_num, url_col)
            cell.hyperlink = first_url
            cell.font = Font(color="0563C1", underline="single")


# ── Single-framework Excel ────────────────────────────────────────────────────

async def _generate_excel(framework_id: str, reports_dir: str, tenant_id: str = None) -> dict:
    from database import get_database
    db = get_database()
    tenant_name = tenant_id or "Unknown Tenant"
    if tenant_id:
        tenant_doc = await db.tenants.find_one({"id": tenant_id})
        tenant_name = tenant_doc.get("name", tenant_id) if tenant_doc else tenant_id

    framework, asset_summary, control_rows = await _build_report_data(framework_id, tenant_id)
    fw_name = framework.get("name", framework_id)
    wb = openpyxl.Workbook()

    # Sheet 1: Asset Summary
    ws1 = wb.active
    ws1.title = "Asset Summary"
    ws1.append([f"Compliance Report: {fw_name}"])
    ws1["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws1.append([f"Tenant: {tenant_name}"])
    ws1.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws1.append([f"Export Date: {datetime.now().strftime('%Y-%m-%d')}"])
    ws1.append([])
    ws1.append(["ASSET COMPLIANCE SUMMARY"])
    ws1[f"A{ws1.max_row}"].fill = _SECTION_FILL
    ws1[f"A{ws1.max_row}"].font = _SECTION_FONT
    ws1.append([])

    if asset_summary:
        headers = list(asset_summary[0].keys())
        _xl_header_row(ws1, headers)
        status_col = headers.index("Overall Status") + 1
        score_col  = headers.index("Score (%)") + 1
        for row in asset_summary:
            ws1.append([_sanitize_cell(v) for v in row.values()])
            r = ws1.max_row
            for c in range(1, len(headers) + 1):
                ws1.cell(r, c).border = _THIN_BORDER
                ws1.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            _apply_status_colors(ws1, r, status_col, row.get("Overall Status", ""))
            ws1.cell(r, score_col).font = Font(bold=True)
    _xl_auto_width(ws1)

    # Sheet 2: Control Details
    ws2 = wb.create_sheet("Control Details")
    ws2.append([f"Control Details with Evidence — {fw_name}"])
    ws2["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws2.append([])

    if control_rows:
        headers2 = list(control_rows[0].keys())
        _xl_header_row(ws2, headers2)
        ctrl_col  = headers2.index("Control Status") + 1
        asset_col = headers2.index("Asset Status") + 1
        url_col = headers2.index("Evidence URLs") + 1 if "Evidence URLs" in headers2 else None
        for row in control_rows:
            ws2.append([_sanitize_cell(v) for v in row.values()])
            r = ws2.max_row
            for c in range(1, len(headers2) + 1):
                ws2.cell(r, c).border = _THIN_BORDER
                ws2.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            _apply_status_colors(ws2, r, ctrl_col, row.get("Control Status", ""))
            _apply_status_colors(ws2, r, asset_col, row.get("Asset Status", ""))
            _apply_url_hyperlink(ws2, r, url_col, row.get("Evidence URLs", "—"))
    _xl_auto_width(ws2)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"compliance_report_{framework_id}_{timestamp}.xlsx"
    wb.save(os.path.join(reports_dir, filename))
    return {"filename": filename, "url": f"/static/reports/{filename}",
            "generatedAt": datetime.now().isoformat(), "rowCount": len(control_rows)}


# ── All-frameworks Excel ───────────────────────────────────────────────────────

async def _generate_all_excel(reports_dir: str, tenant_id: str = None, db=None) -> dict:
    if db is None:
        from database import get_database
        db = get_database()

    tenant_name = tenant_id or "Unknown Tenant"
    if tenant_id:
        tenant_doc = await db.tenants.find_one({"id": tenant_id})
        tenant_name = tenant_doc.get("name", tenant_id) if tenant_doc else tenant_id

    frameworks = await db.compliance_frameworks.find(
        {}, {"id": 1, "name": 1}
    ).to_list(length=100)
    if not frameworks:
        raise ValueError("No compliance frameworks found")

    import logging
    logger = logging.getLogger(__name__)

    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "Overview"
    ov.append(["All Compliance Frameworks — Overview"])
    ov["A1"].font = Font(bold=True, size=14, color="1F3864")
    ov.append([f"Tenant: {tenant_name}"])
    ov.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ov.append([])
    ov.append(["FRAMEWORK SUMMARY"])
    ov[f"A{ov.max_row}"].fill = _SECTION_FILL
    ov[f"A{ov.max_row}"].font = _SECTION_FONT
    ov.append([])

    ov_hdrs = ["Framework", "Total Controls", "Assets Evaluated",
               "Avg Score (%)", "Compliant Assets", "Non-Compliant Assets", "Overall Status"]
    _xl_header_row(ov, ov_hdrs)
    ov_status_col = ov_hdrs.index("Overall Status") + 1

    total_done = 0
    for fw in frameworks:
        fw_id = fw.get("id", "")
        fw_name = fw.get("name", fw_id)
        try:
            _, asset_summary, control_rows = await _build_report_data(fw_id, tenant_id)
        except Exception as exc:
            logger.warning("Skipping framework %s in combined report: %s", fw_name, exc)
            continue
        total_done += 1

        total_controls = len({r["Control ID"] for r in control_rows})
        total_assets = len(asset_summary)
        avg_score = round(
            sum(a["Score (%)"] for a in asset_summary) / total_assets, 1
        ) if total_assets else 0.0
        compliant_ct = sum(1 for a in asset_summary if a["Overall Status"] == "Compliant")
        non_compliant_ct = sum(1 for a in asset_summary if a["Overall Status"] == "Non-Compliant")
        overall = _overall_verdict(avg_score)

        ov.append([fw_name, total_controls, total_assets, avg_score,
                   compliant_ct, non_compliant_ct, overall])
        r = ov.max_row
        for c in range(1, len(ov_hdrs) + 1):
            ov.cell(r, c).border = _THIN_BORDER
            ov.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        _apply_status_colors(ov, r, ov_status_col, overall)

        short = re.sub(r'[\\/?*\[\]:]', '-', fw_name)[:24].strip()

        ws_a = wb.create_sheet(f"{short} Assets")
        ws_a.append([f"Asset Summary — {fw_name}"])
        ws_a["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws_a.append([])
        if asset_summary:
            hdrs_a = list(asset_summary[0].keys())
            _xl_header_row(ws_a, hdrs_a)
            sc_a = hdrs_a.index("Overall Status") + 1
            for row in asset_summary:
                ws_a.append([_sanitize_cell(v) for v in row.values()])
                rn = ws_a.max_row
                for c in range(1, len(hdrs_a) + 1):
                    ws_a.cell(rn, c).border = _THIN_BORDER
                    ws_a.cell(rn, c).alignment = Alignment(vertical="top", wrap_text=True)
                _apply_status_colors(ws_a, rn, sc_a, row.get("Overall Status", ""))
        _xl_auto_width(ws_a)

        ws_c = wb.create_sheet(f"{short} Controls")
        ws_c.append([f"Control Details — {fw_name}"])
        ws_c["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws_c.append([])
        if control_rows:
            hdrs_c = list(control_rows[0].keys())
            _xl_header_row(ws_c, hdrs_c)
            cs_c  = hdrs_c.index("Control Status") + 1
            as_c  = hdrs_c.index("Asset Status") + 1
            url_c = hdrs_c.index("Evidence URLs") + 1 if "Evidence URLs" in hdrs_c else None
            for row in control_rows:
                ws_c.append([_sanitize_cell(v) for v in row.values()])
                rn = ws_c.max_row
                for c in range(1, len(hdrs_c) + 1):
                    ws_c.cell(rn, c).border = _THIN_BORDER
                    ws_c.cell(rn, c).alignment = Alignment(vertical="top", wrap_text=True)
                _apply_status_colors(ws_c, rn, cs_c, row.get("Control Status", ""))
                _apply_status_colors(ws_c, rn, as_c, row.get("Asset Status", ""))
                _apply_url_hyperlink(ws_c, rn, url_c, row.get("Evidence URLs", "—"))
        _xl_auto_width(ws_c)

    _xl_auto_width(ov)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"all_compliance_report_{timestamp}.xlsx"
    wb.save(os.path.join(reports_dir, filename))
    return {"filename": filename, "url": f"/static/reports/{filename}",
            "generatedAt": datetime.now().isoformat(), "frameworkCount": total_done}
