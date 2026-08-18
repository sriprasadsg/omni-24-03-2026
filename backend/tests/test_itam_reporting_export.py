"""ITAM Reporting export/download tests — Phase 72, Plans 01 (tracer) and 05
(PDF/Excel renderers).

Plan 01's tracer proof: a run returning the expiring asset while excluding
an in-warranty and an expired one; a CSV export writing a real file whose
header matches the declared columns and whose data rows equal the full
match set (not the preview page); a zero-row export still writing a valid
header-only file; download of that filename returning 200 for the owning
tenant and 403 for a different tenant; a filename containing traversal
segments returning 400 before any filesystem read; and an unregistered
export format returning 400.

Plan 05 extends this file with the pdf/xlsx renderer behaviors: a real
.pdf/.xlsx file carrying the full match set and declared headers, a
zero-row export that still produces a valid file, a custom saved report
exporting through the same route, tenant-owned download coverage, status
cell coloring, and formula-injection defusal (T-72-06/T-72-11).
"""
import csv
import os
import sys
from datetime import datetime, timedelta, timezone
from unittest.mock import AsyncMock

import openpyxl
import pytest
from fastapi import HTTPException
from httpx import AsyncClient, ASGITransport

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from tests.conftest import make_token_data
from tests.itam_reporting_test_support import (  # noqa: F401 — fixtures re-exported for pytest
    mock_db,
    patch_reporting_get_database,
    reporting_app,
    report_asset,
)

from authentication_service import get_current_user as real_get_current_user
from api_key_auth import get_current_user_or_api_key  # Phase 73 (D-01/D-02): routes gated by _require_itam_admin now resolve through this dependency, not get_current_user
import itam_reporting_endpoints
import itam_reporting_pdf
from itam_reporting_service import _REPORTS_DIR

_WARRANTY_COLUMNS = ["Asset Tag", "Name", "Lifecycle Status", "Warranty Expires", "Days To Expiry", "Status"]


def _no_join_data(mock_db):
    """Seeds every joined collection with an empty result — the shared
    zero-join baseline the custom-report export tests below start from
    (mirrors test_itam_reporting_builder.py's own helper)."""
    mock_db.license_assignments.find.return_value.to_list.return_value = []
    mock_db.components.find.return_value.to_list.return_value = []
    mock_db.itam_consumables.find.return_value.to_list.return_value = []
    mock_db.asset_models.find.return_value.to_list.return_value = []
    mock_db.system_settings.find_one = AsyncMock(return_value=None)


def _client(app, tenant_id="tenant-a", role="admin", username="admin@example.com"):
    current_user = make_token_data(tenant_id=tenant_id, role=role, username=username)
    app.dependency_overrides[get_current_user_or_api_key] = lambda: current_user
    app.dependency_overrides[real_get_current_user] = lambda: current_user
    return AsyncClient(transport=ASGITransport(app=app), base_url="http://testserver")


def _offset_iso(days: int) -> str:
    """purchaseDate that, with warrantyMonths=0 (0 months added is a no-op —
    _add_months returns the purchase date unchanged), makes the warranty
    expire exactly `days` from now — sidesteps calendar-month arithmetic
    entirely for test setup."""
    return (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()


class TestWarrantyExpiringRun:
    """POST /api/itam/reports/prebuilt/warranty_expiring/run"""

    @pytest.mark.asyncio
    async def test_run_returns_expiring_excludes_active_and_expired(self, mock_db, reporting_app):
        expiring = report_asset(
            id="asset-expiring", assetTag="IT-0001",
            purchaseDate=_offset_iso(15), warrantyMonths=0,  # expires in 15 days -> within 30-day default window
        )
        active = report_asset(
            id="asset-active", assetTag="IT-0002",
            purchaseDate=_offset_iso(200), warrantyMonths=0,  # expires in 200 days -> active
        )
        expired = report_asset(
            id="asset-expired", assetTag="IT-0003",
            purchaseDate=_offset_iso(-10), warrantyMonths=0,  # expired 10 days ago
        )
        mock_db.assets.find.return_value.to_list.return_value = [expiring, active, expired]
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/run")

        assert r.status_code == 200, r.text
        body = r.json()
        assert [row["Asset Tag"] for row in body["rows"]] == ["IT-0001"]
        assert body["rowCount"] == 1
        assert body["columns"] == _WARRANTY_COLUMNS
        assert body["truncated"] is False

    @pytest.mark.asyncio
    async def test_run_unknown_report_key_returns_404(self, mock_db, reporting_app):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)
        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/not_a_real_report/run")
        assert r.status_code == 404


class TestWarrantyExpiringExport:
    """POST /api/itam/reports/prebuilt/warranty_expiring/export"""

    @pytest.mark.asyncio
    async def test_csv_export_writes_full_match_set_not_just_preview_page(self, mock_db, reporting_app):
        assets = [
            report_asset(id=f"asset-{i}", assetTag=f"IT-{i:04d}", purchaseDate=_offset_iso(15), warrantyMonths=0)
            for i in range(3)
        ]
        mock_db.assets.find.return_value.to_list.return_value = assets
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=csv")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 3
        assert body["truncated"] is False

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        assert os.path.isfile(file_path)
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = list(csv.reader(f))
        assert reader[0] == _WARRANTY_COLUMNS
        assert len(reader) - 1 == 3  # header + all 3 matches, not a paged subset

    @pytest.mark.asyncio
    async def test_zero_row_export_writes_header_only_file(self, mock_db, reporting_app):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=csv")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 0

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = list(csv.reader(f))
        assert len(reader) == 1  # exactly one header line, zero data lines — never an error
        assert reader[0] == _WARRANTY_COLUMNS

    @pytest.mark.asyncio
    async def test_unregistered_format_returns_400(self, mock_db, reporting_app):
        # "xml" (not "pdf") — Plan 72-05 registers pdf/xlsx into RENDERERS,
        # so a genuinely unregistered format is needed to exercise this
        # 400 path now that pdf is a real, activated format.
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)
        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=xml")
        assert r.status_code == 400
        assert "pdf" in r.json()["detail"]


class TestReportDownload:
    """GET /api/itam/reports/download/{filename}"""

    async def _export_and_get_filename(self, reporting_app):
        async with _client(reporting_app, tenant_id="tenant-a") as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=csv")
        assert r.status_code == 200, r.text
        return r.json()["filename"]

    @pytest.mark.asyncio
    async def test_download_owning_tenant_returns_200(self, mock_db, reporting_app):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)
        filename = await self._export_and_get_filename(reporting_app)
        mock_db.itam_report_exports.find_one = AsyncMock(
            return_value={"filename": filename, "tenantId": "tenant-a"}
        )

        async with _client(reporting_app, tenant_id="tenant-a") as ac:
            r = await ac.get(f"/api/itam/reports/download/{filename}")
        assert r.status_code == 200, r.text

    @pytest.mark.asyncio
    async def test_download_cross_tenant_returns_403(self, mock_db, reporting_app, patch_reporting_get_database):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)
        filename = await self._export_and_get_filename(reporting_app)
        # The export was persisted under tenant-a; a tenant-b caller must be
        # rejected even though the filename is exact and correct.
        mock_db.itam_report_exports.find_one = AsyncMock(
            return_value={"filename": filename, "tenantId": "tenant-a"}
        )

        patch_reporting_get_database("tenant-b")
        async with _client(reporting_app, tenant_id="tenant-b") as ac:
            r = await ac.get(f"/api/itam/reports/download/{filename}")
        assert r.status_code == 403


class TestPathTraversalGuard:
    """T-72-03: a filename whose resolved path escapes the reports directory
    is rejected with 400 before any filesystem read. Called directly against
    the route function (bypassing ASGI routing) because a URL-encoded '/'
    segment is normalised away by Starlette's own router before reaching
    application code, which would 404 rather than exercise this module's
    containment check — the defense under test."""

    @pytest.mark.asyncio
    async def test_traversal_filename_rejected_with_400(self):
        admin = make_token_data(tenant_id="tenant-a", role="admin", username="admin@example.com")
        with pytest.raises(HTTPException) as exc_info:
            await itam_reporting_endpoints.download_report(
                filename="../../../../etc/passwd", current_user=admin,
            )
        assert exc_info.value.status_code == 400


# ── Plan 05: PDF renderer (Task 1) ──────────────────────────────────────────

class TestWarrantyExpiringPdfExport:
    """POST /api/itam/reports/prebuilt/warranty_expiring/export?format=pdf"""

    @pytest.mark.asyncio
    async def test_pdf_export_writes_real_pdf_carrying_full_match_set(self, mock_db, reporting_app):
        # 60 rows > the default 50-row preview page, proving the export
        # carries the full match count rather than one preview page.
        assets = [
            report_asset(id=f"asset-{i}", assetTag=f"IT-{i:04d}", purchaseDate=_offset_iso(15), warrantyMonths=0)
            for i in range(60)
        ]
        mock_db.assets.find.return_value.to_list.return_value = assets
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=pdf")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 60
        assert body["filename"].endswith(".pdf")

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        assert os.path.isfile(file_path)
        with open(file_path, "rb") as f:
            magic = f.read(4)
        assert magic == b"%PDF"

    @pytest.mark.asyncio
    async def test_zero_row_pdf_export_returns_200_with_real_file(self, mock_db, reporting_app):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=pdf")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 0

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        with open(file_path, "rb") as f:
            magic = f.read(4)
        assert magic == b"%PDF"
        assert os.path.getsize(file_path) > 0

    @pytest.mark.asyncio
    async def test_pdf_export_recorded_in_exports_and_cross_tenant_403(
        self, mock_db, reporting_app, patch_reporting_get_database,
    ):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app, tenant_id="tenant-a") as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=pdf")
        assert r.status_code == 200, r.text
        filename = r.json()["filename"]

        mock_db.itam_report_exports.find_one = AsyncMock(
            return_value={"filename": filename, "tenantId": "tenant-a"}
        )

        # Owning tenant can download.
        async with _client(reporting_app, tenant_id="tenant-a") as ac:
            r = await ac.get(f"/api/itam/reports/download/{filename}")
        assert r.status_code == 200, r.text

        # A different tenant is rejected even with the exact correct filename.
        patch_reporting_get_database("tenant-b")
        async with _client(reporting_app, tenant_id="tenant-b") as ac:
            r = await ac.get(f"/api/itam/reports/download/{filename}")
        assert r.status_code == 403


class TestPdfCustomReportExport:
    """POST /api/itam/reports/custom/{report_id}/export?format=pdf"""

    @pytest.mark.asyncio
    async def test_custom_report_exports_to_pdf_through_same_route(self, mock_db, reporting_app):
        saved_doc = {
            "id": "rpt-pdf-1", "tenantId": "tenant-a", "name": "PDF Custom Report",
            "columns": ["asset.assetTag", "asset.name"], "filters": [],
        }
        mock_db.itam_reports.find_one = AsyncMock(return_value=saved_doc)
        assets = [report_asset(id=f"a{i}", assetTag=f"IT-{i:04d}") for i in range(5)]
        mock_db.assets.find.return_value.to_list.return_value = assets
        _no_join_data(mock_db)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/custom/rpt-pdf-1/export?format=pdf")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 5

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        with open(file_path, "rb") as f:
            magic = f.read(4)
        assert magic == b"%PDF"


class TestPdfFormulaInjectionDefused:
    """T-72-06/T-72-11: every cell passes through _sanitize_cell before it
    reaches the PDF, in addition to the html.escape applied before wrapping
    in a reportlab Paragraph."""

    @pytest.mark.asyncio
    async def test_formula_trigger_cell_is_sanitized_before_rendering(self, monkeypatch, tmp_path):
        calls = []
        real_sanitize = itam_reporting_pdf._sanitize_cell

        def _spy_sanitize(v):
            calls.append(v)
            return real_sanitize(v)

        monkeypatch.setattr(itam_reporting_pdf, "_sanitize_cell", _spy_sanitize)

        report = {
            "key": "warranty_expiring",
            "title": "Warranty Expiring",
            "columns": ["Asset Tag", "Name", "Status"],
            "rows": [{"Asset Tag": "IT-0001", "Name": "=SUM(A1:A10)", "Status": "Expiring"}],
            "rowCount": 1,
            "truncated": False,
        }
        result = await itam_reporting_pdf._generate_pdf(report, str(tmp_path), "tenant-a")

        assert "=SUM(A1:A10)" in calls
        file_path = os.path.join(str(tmp_path), result["filename"])
        assert os.path.isfile(file_path)


# ── Plan 05: Excel renderer (Task 2) ────────────────────────────────────────

class TestWarrantyExpiringExcelExport:
    """POST /api/itam/reports/prebuilt/warranty_expiring/export?format=xlsx"""

    @pytest.mark.asyncio
    async def test_xlsx_export_header_and_full_match_set(self, mock_db, reporting_app):
        assets = [
            report_asset(id=f"asset-{i}", assetTag=f"IT-{i:04d}", purchaseDate=_offset_iso(15), warrantyMonths=0)
            for i in range(60)
        ]
        mock_db.assets.find.return_value.to_list.return_value = assets
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=xlsx")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 60
        assert body["filename"].endswith(".xlsx")

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == _WARRANTY_COLUMNS
        assert ws.max_row - 1 == 60  # header + all 60 matches, not a paged subset

    @pytest.mark.asyncio
    async def test_zero_row_xlsx_export_has_header_only(self, mock_db, reporting_app):
        mock_db.assets.find.return_value.to_list.return_value = []
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=xlsx")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 0

        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        assert ws.max_row == 1
        assert [c.value for c in ws[1]] == _WARRANTY_COLUMNS

    @pytest.mark.asyncio
    async def test_status_cell_receives_fill_for_expiring(self, mock_db, reporting_app):
        assets = [
            report_asset(id="asset-1", assetTag="IT-0001", purchaseDate=_offset_iso(15), warrantyMonths=0),
        ]
        mock_db.assets.find.return_value.to_list.return_value = assets
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/prebuilt/warranty_expiring/export?format=xlsx")

        assert r.status_code == 200, r.text
        body = r.json()
        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        status_col = _WARRANTY_COLUMNS.index("Status") + 1
        # itam_finance_service.WARRANTY_STATUS_EXPIRING is lowercase
        # ("expiring") — the color-table lookup must match it
        # case-insensitively rather than only the Title-case literal.
        assert ws.cell(2, status_col).value == "expiring"
        fill_rgb = ws.cell(2, status_col).fill.fgColor.rgb or ""
        assert "FFEB9C" in fill_rgb


class TestXlsxCustomReportExport:
    """POST /api/itam/reports/custom/{report_id}/export?format=xlsx"""

    @pytest.mark.asyncio
    async def test_custom_report_exports_to_xlsx_through_same_route(self, mock_db, reporting_app):
        saved_doc = {
            "id": "rpt-xlsx-1", "tenantId": "tenant-a", "name": "Excel Custom Report",
            "columns": ["asset.assetTag", "asset.name"], "filters": [],
        }
        mock_db.itam_reports.find_one = AsyncMock(return_value=saved_doc)
        assets = [report_asset(id=f"a{i}", assetTag=f"IT-{i:04d}") for i in range(5)]
        mock_db.assets.find.return_value.to_list.return_value = assets
        _no_join_data(mock_db)

        async with _client(reporting_app) as ac:
            r = await ac.post("/api/itam/reports/custom/rpt-xlsx-1/export?format=xlsx")

        assert r.status_code == 200, r.text
        body = r.json()
        assert body["rowCount"] == 5
        file_path = os.path.join(_REPORTS_DIR, body["filename"])
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        assert ws.max_row - 1 == 5


class TestXlsxFormulaInjectionDefused:
    """T-72-06: every cell passes through _sanitize_cell before it reaches
    the workbook — asserted by reading the actual written cell value."""

    @pytest.mark.asyncio
    async def test_formula_trigger_cell_is_written_defused(self, tmp_path):
        import itam_reporting_excel

        report = {
            "key": "warranty_expiring",
            "title": "Warranty Expiring",
            "columns": ["Asset Tag", "Name", "Status"],
            "rows": [{"Asset Tag": "IT-0001", "Name": "=SUM(A1:A10)", "Status": "Expiring"}],
            "rowCount": 1,
            "truncated": False,
        }
        result = await itam_reporting_excel._generate_excel(report, str(tmp_path), "tenant-a")

        file_path = os.path.join(str(tmp_path), result["filename"])
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        name_col = ["Asset Tag", "Name", "Status"].index("Name") + 1
        assert ws.cell(2, name_col).value == "'=SUM(A1:A10)"


class TestExportFormatsAgree:
    """D-11/D-12/D-14: the same report exported in all three formats reports
    an identical rowCount, since all three read the same shared row set
    through build_report_rows."""

    @pytest.mark.asyncio
    async def test_csv_pdf_xlsx_report_identical_row_count(self, mock_db, reporting_app):
        assets = [
            report_asset(id=f"asset-{i}", assetTag=f"IT-{i:04d}", purchaseDate=_offset_iso(15), warrantyMonths=0)
            for i in range(7)
        ]
        mock_db.assets.find.return_value.to_list.return_value = assets
        mock_db.system_settings.find_one = AsyncMock(return_value=None)

        row_counts = {}
        for fmt in ("csv", "pdf", "xlsx"):
            async with _client(reporting_app) as ac:
                r = await ac.post(f"/api/itam/reports/prebuilt/warranty_expiring/export?format={fmt}")
            assert r.status_code == 200, r.text
            row_counts[fmt] = r.json()["rowCount"]

        assert row_counts["csv"] == row_counts["pdf"] == row_counts["xlsx"] == 7
