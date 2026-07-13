from datetime import datetime, timezone
import csv
import io
import logging
from typing import Any, Dict, List, Optional
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel

from database import get_database
from tenant_context import get_tenant_id

logger = logging.getLogger(__name__)

MAX_QUESTION_TEXT_LENGTH = 4000
ACCEPTED_QUESTION_HEADERS = ["question", "question text", "control question"]


class QuestionnaireInboundCreate(BaseModel):
    title: str
    questions: Optional[List[Dict[str, Any]]] = None


class QuestionnaireInboundService:
    def __init__(self):
        self._db = get_database

    async def create_question_set(
        self, data: QuestionnaireInboundCreate, tenant_id: str, created_by: str
    ) -> Dict[str, Any]:
        doc = {
            "id": f"qni-{uuid4().hex}",
            "tenantId": tenant_id,
            "title": data.title,
            "questions": data.questions if data.questions is not None else [],
            "status": "Draft",
            "createdBy": created_by,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await self._db().questionnaire_inbound.insert_one(doc)
        doc.pop("_id", None)
        return doc

    async def list_question_sets(self, tenant_id: str) -> List[Dict[str, Any]]:
        cursor = self._db().questionnaire_inbound.find({"tenantId": tenant_id}, {"_id": 0})
        return await cursor.to_list(length=None)

    async def get_question_set(self, qid: str, tenant_id: str) -> Optional[Dict[str, Any]]:
        return await self._db().questionnaire_inbound.find_one({"id": qid, "tenantId": tenant_id}, {"_id": 0})

    async def parse_upload(self, filename: str, content_bytes: bytes) -> List[Dict[str, Any]]:
        questions: List[Dict[str, Any]] = []
        if filename.endswith(".csv"):
            decoded_content = content_bytes.decode("utf-8")
            reader = csv.DictReader(io.StringIO(decoded_content))
            fieldnames = reader.fieldnames or []
            headers = [h.strip().lower() for h in fieldnames]
            question_col = self._find_question_column(headers)
            # DictReader rows are keyed by the ORIGINAL header, not the
            # normalized one used for matching
            original_col = fieldnames[headers.index(question_col)]
            for i, row in enumerate(reader):
                text = self._sanitize_question_text(row[original_col])
                if text:
                    questions.append({"id": uuid4().hex, "text": text, "rowIndex": i})
        elif filename.endswith((".xlsx", ".xls")):
            try:
                workbook = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
                sheet = workbook.active
                headers = [str(cell.value).strip().lower() for cell in sheet[1]]
                question_col_idx = self._find_question_column_index(headers)
                for i, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    text = self._sanitize_question_text(row_data[question_col_idx])
                    if text:
                        questions.append({"id": uuid4().hex, "text": text, "rowIndex": i + 1})
            except ImportError:
                raise HTTPException(
                    status_code=422,
                    detail="openpyxl not installed. Please install it to process Excel files.",
                )
            except Exception as e:
                logger.error(f"Error parsing Excel file {filename}: {e}")
                raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {e}")
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")

        return questions

    def _find_question_column(self, headers: List[str]) -> str:
        for header in headers:
            if header in ACCEPTED_QUESTION_HEADERS:
                return header
        raise HTTPException(
            status_code=400,
            detail=f"No recognized question column header found. Accepted headers: {', '.join(ACCEPTED_QUESTION_HEADERS)}",
        )

    def _find_question_column_index(self, headers: List[str]) -> int:
        for i, header in enumerate(headers):
            if header in ACCEPTED_QUESTION_HEADERS:
                return i
        raise HTTPException(
            status_code=400,
            detail=f"No recognized question column header found. Accepted headers: {', '.join(ACCEPTED_QUESTION_HEADERS)}",
        )

    def _sanitize_question_text(self, text: Any) -> str:
        if text is None:
            return ""
        text = str(text).strip()
        # Strip control characters (remove anything not printable ASCII or common Unicode)
        text = "".join(c for c in text if c.isprintable() or c in ("\n", "\t"))
        return text[:MAX_QUESTION_TEXT_LENGTH]


questionnaire_inbound_service = QuestionnaireInboundService()
