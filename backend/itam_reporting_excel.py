"""ITAM reporting: Excel (XLSX) renderer (Phase 72, Plan 05, D-11).

Reuses compliance_reporting_excel.py's domain-agnostic worksheet helpers
(_xl_header_row, _xl_auto_width) directly — they operate purely on
headers/values with no compliance-specific assumption. The status-fill
table itself IS domain-specific (compliance_reporting_excel._STATUS_FILLS
is keyed by Compliant/Non-Compliant/... and never matches an ITAM report's
Status column), so this module defines its own ITAM status-color table and
a local _apply_status_colors following the exact same
PatternFill+Font-lookup pattern rather than inventing a third styling
system. Registers itself into itam_reporting_service.RENDERERS under the
"xlsx" key.
"""

import os
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from compliance_reporting_data import _sanitize_cell
from compliance_reporting_excel import _xl_auto_width, _xl_header_row

_SHEET_NAME_LIMIT = 31  # Excel/OOXML sheet-title cap.

# ITAM status vocabulary these reports actually emit (warranty_expiring,
# license_utilization, overdue_audits, low_stock_consumables). Keyed
# lowercase and matched case-insensitively below:
# itam_finance_service.WARRANTY_STATUS_EXPIRING/_EXPIRED/_ACTIVE are
# lowercase ("expiring"/"expired"/"active"), while
# itam_reporting_prebuilt._build_license_utilization_rows emits Title-case
# ("Active"/"Expired") for the same concept — a pre-existing case
# inconsistency across reports this renderer must tolerate rather than
# silently under-color.
_XL_STATUS_FILLS = {
    "active":        PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "expiring":      PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "expired":       PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "low stock":     PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "overdue":       PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "never audited": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
_XL_STATUS_FONTS = {
    "active":        Font(color="006100"),
    "expiring":      Font(color="9C5700"),
    "expired":       Font(color="9C0006"),
    "low stock":     Font(color="9C5700"),
    "overdue":       Font(color="9C0006"),
    "never audited": Font(color="9C0006"),
}


def _apply_status_colors(ws, row_num, col_idx, val):
    key = str(val).strip().lower() if val is not None else ""
    if _XL_STATUS_FILLS.get(key):
        ws.cell(row_num, col_idx).fill = _XL_STATUS_FILLS[key]
    if _XL_STATUS_FONTS.get(key):
        ws.cell(row_num, col_idx).font = _XL_STATUS_FONTS[key]


def _status_column_index(columns):
    """Whichever column a report declares as its status column — the
    literal 'Status' header, when present."""
    return columns.index("Status") if "Status" in columns else None


async def _generate_excel(report: dict, reports_dir: str, tenant_id: str = None) -> dict:
    """Writes report['rows'] (a list of {column: value} dicts keyed by
    report['columns']) into one worksheet titled from report['title'].
    Every cell passes through _sanitize_cell (CWE-1236 formula-injection
    defence, T-72-06). A zero-row report still produces a workbook with
    only the header row."""
    columns = report["columns"]
    rows = report["rows"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (str(report["title"])[:_SHEET_NAME_LIMIT] or "Report")

    _xl_header_row(ws, columns)
    status_idx = _status_column_index(columns)
    status_col_1based = status_idx + 1 if status_idx is not None else None

    for row in rows:
        values = [_sanitize_cell(row.get(col)) for col in columns]
        ws.append(values)
        r = ws.max_row
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        if status_col_1based is not None:
            _apply_status_colors(ws, r, status_col_1based, row.get("Status"))

    if report.get("truncated"):
        ws.append([
            "TRUNCATED — the export ceiling was reached; not every "
            "matching row is shown in this file."
        ])

    _xl_auto_width(ws)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"itam_report_{report['key']}_{timestamp}.xlsx"
    wb.save(os.path.join(reports_dir, filename))

    return {
        "filename": filename,
        "url": f"/static/reports/{filename}",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "rowCount": report["rowCount"],
        "truncated": report.get("truncated", False),
    }
